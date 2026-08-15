from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils.text import slugify
from django.db.models import Sum, Count, Q
from django.db.models.functions import ExtractMonth
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
import json
import csv
from datetime import datetime
from django.core.files.storage import FileSystemStorage

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
import os
try:
    from svglib.svglib import svg2rlg
    from reportlab.graphics import renderPDF
    _SVGLIB_OK = True
except ImportError:
    _SVGLIB_OK = False

from django.contrib.auth.models import User
from .models import Category, Service, Realization, Product, Order, OrderItem, ContactMessage


# ─── Garde : superusers uniquement ──────────────────────────
def is_superuser(user):
    return user.is_authenticated and user.is_superuser


superuser_required = user_passes_test(is_superuser, login_url='/dashboard/login/')


# ─── Login / Logout personnalisés ───────────────────────────
def dashboard_login(request):
    """Page de connexion réservée aux superusers."""
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('dashboard_home')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is None and username:
            user_obj = User.objects.filter(username__iexact=username).first()
            if user_obj:
                user = authenticate(request, username=user_obj.username, password=password)

        if user is not None:
            if user.is_superuser:
                login(request, user)
                next_url = request.GET.get('next') or request.POST.get('next') or 'dashboard_home'
                return redirect(next_url)
            else:
                error = "⛔ Accès refusé. Seuls les super-administrateurs peuvent accéder au dashboard."
        else:
            error = "❌ Identifiants incorrects. Veuillez réessayer."

    return render(request, 'admin_custom/login.html', {'error': error})


def dashboard_logout(request):
    """Déconnexion et retour à la page de login."""
    logout(request)
    return redirect('dashboard_login')


# ============================================================
# DASHBOARD HOME
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_home(request):
    # Stats
    total_products = Product.objects.count()
    active_products = Product.objects.filter(is_active=True).count()
    total_services = Service.objects.count()
    total_realizations = Realization.objects.count()
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    unread_messages = ContactMessage.objects.filter(is_read=False).count()
    total_revenue = Order.objects.filter(status__in=['confirmed', 'shipped', 'delivered']).aggregate(
        total=Sum('total_amount'))['total'] or 0

    recent_orders = Order.objects.all()[:5]
    recent_messages = ContactMessage.objects.filter(is_read=False)[:5]
    low_stock_products = Product.objects.filter(is_active=True, stock__lte=3)

    # Statistiques mensuelles pour Chart.js (6 derniers mois)
    sales_by_month = (
        Order.objects.filter(status__in=['confirmed', 'shipped', 'delivered'])
        .annotate(month=ExtractMonth('created_at'))
        .values('month')
        .annotate(total=Sum('total_amount'), count=Count('id'))
        .order_by('month')
    )
    
    # Remplir les données par défaut pour les 6 derniers mois
    month_names = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jui", "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]
    current_month = datetime.now().month
    chart_labels = []
    chart_revenue = []
    chart_count = []
    
    for i in range(5, -1, -1):
        m = (current_month - i - 1) % 12 + 1
        chart_labels.append(month_names[m - 1])
        
        # Trouver la valeur correspondante dans la requête
        matching = next((x for x in sales_by_month if x['month'] == m), None)
        chart_revenue.append(float(matching['total']) if matching and matching['total'] else 0.0)
        chart_count.append(matching['count'] if matching else 0)

    context = {
        'total_products': total_products,
        'active_products': active_products,
        'total_services': total_services,
        'total_realizations': total_realizations,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'unread_messages': unread_messages,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
        'recent_messages': recent_messages,
        'low_stock_products': low_stock_products,
        'chart_labels': json.dumps(chart_labels),
        'chart_revenue': json.dumps(chart_revenue),
        'chart_count': json.dumps(chart_count),
        'page_title': 'Tableau de Bord',
    }
    return render(request, 'admin_custom/dashboard.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_export_orders(request):
    """Exporte les commandes au format CSV."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="Rapport_Commandes_STANTECH_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['N° Commande', 'Client', 'Email', 'Téléphone', 'Adresse', 'Ville', 'Montant (FCFA)', 'Statut', 'Date'])

    orders = Order.objects.all().order_by('-created_at')
    status_labels = dict(Order.STATUS_CHOICES)

    for o in orders:
        writer.writerow([
            o.order_number,
            o.customer_name,
            o.customer_email,
            o.customer_phone,
            o.delivery_address,
            o.city,
            int(o.total_amount),
            status_labels.get(o.status, o.status),
            o.created_at.strftime('%d/%m/%Y %H:%M')
        ])

    return response


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_export_orders_excel(request):
    """Exporte le rapport des commandes au format Excel (.xlsx) stylisé."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes STANTECH"

    # En-tête principal
    ws.merge_cells('A1:I1')
    header_cell = ws['A1']
    header_cell.value = "STANTECH ENTERPRISE - RAPPORT DE GESTION DES COMMANDES"
    header_cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Date d'export
    ws.merge_cells('A2:I2')
    sub_cell = ws['A2']
    sub_cell.value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    sub_cell.font = Font(name='Calibri', size=10, italic=True, color="475569")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = ['N° Commande', 'Client', 'Email', 'Téléphone', 'Adresse', 'Ville', 'Montant (FCFA)', 'Statut', 'Date']
    ws.append([]) # Ligne 3 vide

    ws.append(headers)
    ws.row_dimensions[4].height = 25
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    orders = Order.objects.all().order_by('-created_at')
    status_labels = dict(Order.STATUS_CHOICES)

    row_idx = 5
    total_revenue = 0

    for order in orders:
        total_revenue += float(order.total_amount)
        row_data = [
            order.order_number,
            order.customer_name,
            order.customer_email or '—',
            order.customer_phone,
            order.delivery_address,
            order.city,
            float(order.total_amount),
            status_labels.get(order.status, order.status),
            order.created_at.strftime('%d/%m/%Y %H:%M')
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 22

        for col_num in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            if col_num in [1, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if col_num == 7:
                cell.number_format = '#,##0 "FCFA"'

        row_idx += 1

    # Ligne de Total
    ws.append([])
    ws.cell(row=row_idx, column=6, value="TOTAL GENERAL :").font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
    
    total_cell = ws.cell(row=row_idx, column=7, value=total_revenue)
    total_cell.font = Font(name='Calibri', size=11, bold=True, color="059669")
    total_cell.number_format = '#,##0 "FCFA"'
    total_cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    ws.row_dimensions[row_idx].height = 25

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row in [1, 2]: continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Rapport_Commandes_STANTECH_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_export_orders_pdf(request):
    """Exporte le rapport global des commandes au format PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#059669'),
        spaceAfter=4
    )
    sub_style = ParagraphStyle(
        'SubStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=15
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#1E293B')
    )
    cell_header = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )

    # ── En-tête avec logo ──
    _svg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'images', 'stantech_logo.svg')
    if _SVGLIB_OK and os.path.exists(_svg_path):
        drawing = svg2rlg(_svg_path)
        if drawing:
            sx = 180 / drawing.width
            drawing.width = 180
            drawing.height = drawing.height * sx
            drawing.transform = (sx, 0, 0, sx, 0, 0)
            story.append(drawing)
            story.append(Spacer(1, 4))
        else:
            story.append(Paragraph("STANTECH ENTERPRISE", title_style))
    else:
        story.append(Paragraph("STANTECH ENTERPRISE", title_style))
    story.append(Paragraph(f"Rapport des Commandes Client — Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E2E8F0'), spaceAfter=15))

    orders = Order.objects.all().order_by('-created_at')
    status_labels = dict(Order.STATUS_CHOICES)

    table_data = [
        [
            Paragraph("N° Commande", cell_header),
            Paragraph("Client", cell_header),
            Paragraph("Téléphone", cell_header),
            Paragraph("Ville", cell_header),
            Paragraph("Montant", cell_header),
            Paragraph("Statut", cell_header),
            Paragraph("Date", cell_header)
        ]
    ]

    total_amount = 0

    for o in orders:
        total_amount += o.total_amount
        table_data.append([
            Paragraph(f"<b>#{o.order_number}</b>", cell_style),
            Paragraph(o.customer_name, cell_style),
            Paragraph(o.customer_phone, cell_style),
            Paragraph(o.city, cell_style),
            Paragraph(f"<b>{int(o.total_amount):,} FCFA</b>".replace(',', ' '), cell_style),
            Paragraph(status_labels.get(o.status, o.status), cell_style),
            Paragraph(o.created_at.strftime('%d/%m/%Y'), cell_style)
        ])

    t = Table(table_data, colWidths=[90, 100, 80, 70, 85, 60, 50])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#059669')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
    ]))

    story.append(t)
    story.append(Spacer(1, 15))

    summary_text = f"<b>Total Commandes :</b> {orders.count()} &nbsp;&nbsp;|&nbsp;&nbsp; <b>Chiffre d'Affaires Total :</b> {int(total_amount):,} FCFA".replace(',', ' ')
    story.append(Paragraph(summary_text, ParagraphStyle('SumStyle', parent=styles['Normal'], fontSize=11, textColor=colors.HexColor('#059669'), fontName='Helvetica-Bold')))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    filename = f"Rapport_Commandes_STANTECH_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_export_order_single_pdf(request, pk):
    """Génère la facture / bon de commande PDF pour une commande spécifique."""
    order = get_object_or_404(Order, pk=pk)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=35, leftMargin=35, topMargin=35, bottomMargin=35)
    story = []

    styles = getSampleStyleSheet()

    header_brand = ParagraphStyle('Brand', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, textColor=colors.HexColor('#059669'))
    sub_title = ParagraphStyle('Sub', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#64748B'))
    doc_type = ParagraphStyle('DocType', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=14, textColor=colors.HexColor('#1E293B'), alignment=2)
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#475569'))
    val_style = ParagraphStyle('Val', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#1E293B'))

    _svg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'static', 'images', 'stantech_logo.svg')
    logo_cell = None
    if _SVGLIB_OK and os.path.exists(_svg_path):
        _drawing = svg2rlg(_svg_path)
        if _drawing:
            _sx = 160 / _drawing.width
            _drawing.width = 160
            _drawing.height = _drawing.height * _sx
            _drawing.transform = (_sx, 0, 0, _sx, 0, 0)
            logo_cell = _drawing
    if logo_cell is None:
        logo_cell = Paragraph("STANTECH ENTERPRISE<br/><font size=8 color='#64748B'>Solutions Technologiques &amp; Ingénierie<br/>Douala / Abidjan | +237 674861509</font>", header_brand)

    header_data = [
        [
            logo_cell,
            Paragraph(f"FACTURE / BON<br/><font size=10 color='#059669'>#{order.order_number}</font><br/><font size=8 color='#64748B'>Date: {order.created_at.strftime('%d/%m/%Y %H:%M')}</font>", doc_type)
        ]
    ]
    t_header = Table(header_data, colWidths=[320, 200])
    t_header.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (1,0), (1,0), 'RIGHT')
    ]))
    story.append(t_header)
    story.append(Spacer(1, 15))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#059669'), spaceAfter=15))

    client_info = [
        [Paragraph("CLIENT :", label_style), Paragraph(order.customer_name, val_style), Paragraph("TÉLÉPHONE :", label_style), Paragraph(order.customer_phone, val_style)],
        [Paragraph("EMAIL :", label_style), Paragraph(order.customer_email or "—", val_style), Paragraph("VILLE :", label_style), Paragraph(order.city, val_style)],
        [Paragraph("ADRESSE :", label_style), Paragraph(order.delivery_address, val_style), Paragraph("STATUT :", label_style), Paragraph(f"<b>{order.get_status_display()}</b>", val_style)]
    ]
    t_client = Table(client_info, colWidths=[70, 190, 80, 180])
    t_client.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('RIGHTPADDING', (0,0), (-1,-1), 8),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0'))
    ]))
    story.append(t_client)
    story.append(Spacer(1, 20))

    cell_head = ParagraphStyle('Head', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white)
    cell_item = ParagraphStyle('Item', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#1E293B'))

    items_data = [
        [Paragraph("Produit / Article", cell_head), Paragraph("Prix Unitaire", cell_head), Paragraph("Qté", cell_head), Paragraph("Sous-Total", cell_head)]
    ]

    for item in order.items.all():
        items_data.append([
            Paragraph(item.product_name, cell_item),
            Paragraph(f"{int(item.price):,} FCFA".replace(',', ' '), cell_item),
            Paragraph(str(item.quantity), cell_item),
            Paragraph(f"<b>{int(item.subtotal):,} FCFA</b>".replace(',', ' '), cell_item)
        ])

    t_items = Table(items_data, colWidths=[240, 100, 50, 130])
    t_items.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('ALIGN', (2,0), (2,-1), 'CENTER'),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ]))
    story.append(t_items)
    story.append(Spacer(1, 15))

    total_text = f"<b>TOTAL A PAYER :</b> &nbsp;&nbsp; <font size=13 color='#059669'>{int(order.total_amount):,} FCFA</font>".replace(',', ' ')
    p_total = Paragraph(total_text, ParagraphStyle('Tot', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, alignment=2))
    story.append(p_total)

    if order.notes:
        story.append(Spacer(1, 15))
        story.append(Paragraph(f"<b>Notes & Instructions :</b> {order.notes}", val_style))

    story.append(Spacer(1, 30))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#CBD5E1'), spaceAfter=10))
    story.append(Paragraph("Merci de votre confiance. STANTECH — Solutions Technologiques de Pointe.", ParagraphStyle('Foot', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=8, textColor=colors.HexColor('#64748B'), alignment=1)))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Facture_STANTECH_{order.order_number}.pdf"'
    response.write(pdf)
    return response


# ─── Garde : superusers uniquement ──────────────────────────
def is_superuser(user):
    return user.is_authenticated and user.is_superuser


superuser_required = user_passes_test(is_superuser, login_url='/dashboard/login/')


# ─── Login / Logout personnalisés ───────────────────────────
def dashboard_login(request):
    """Page de connexion réservée aux superusers."""
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('dashboard_home')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is None and username:
            user_obj = User.objects.filter(username__iexact=username).first()
            if user_obj:
                user = authenticate(request, username=user_obj.username, password=password)

        if user is not None:
            if user.is_superuser:
                login(request, user)
                next_url = request.GET.get('next') or request.POST.get('next') or 'dashboard_home'
                return redirect(next_url)
            else:
                error = "⛔ Accès refusé. Seuls les super-administrateurs peuvent accéder au dashboard."
        else:
            error = "❌ Identifiants incorrects. Veuillez réessayer."

    return render(request, 'admin_custom/login.html', {'error': error})


def dashboard_logout(request):
    """Déconnexion et retour à la page de login."""
    logout(request)
    return redirect('dashboard_login')


# ============================================================
# DASHBOARD HOME
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_home(request):
    # Stats
    total_products = Product.objects.count()
    active_products = Product.objects.filter(is_active=True).count()
    total_services = Service.objects.count()
    total_realizations = Realization.objects.count()
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    unread_messages = ContactMessage.objects.filter(is_read=False).count()
    total_revenue = Order.objects.filter(status__in=['confirmed', 'shipped', 'delivered']).aggregate(
        total=Sum('total_amount'))['total'] or 0

    recent_orders = Order.objects.all()[:5]
    recent_messages = ContactMessage.objects.filter(is_read=False)[:5]
    low_stock_products = Product.objects.filter(is_active=True, stock__lte=3)

    # Statistiques mensuelles pour Chart.js (6 derniers mois)
    sales_by_month = (
        Order.objects.filter(status__in=['confirmed', 'shipped', 'delivered'])
        .annotate(month=ExtractMonth('created_at'))
        .values('month')
        .annotate(total=Sum('total_amount'), count=Count('id'))
        .order_by('month')
    )
    
    # Remplir les données par défaut pour les 6 derniers mois
    month_names = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jui", "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]
    current_month = datetime.now().month
    chart_labels = []
    chart_revenue = []
    chart_count = []
    
    for i in range(5, -1, -1):
        m = (current_month - i - 1) % 12 + 1
        chart_labels.append(month_names[m - 1])
        
        # Trouver la valeur correspondante dans la requête
        matching = next((x for x in sales_by_month if x['month'] == m), None)
        chart_revenue.append(float(matching['total']) if matching and matching['total'] else 0.0)
        chart_count.append(matching['count'] if matching else 0)

    context = {
        'total_products': total_products,
        'active_products': active_products,
        'total_services': total_services,
        'total_realizations': total_realizations,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'unread_messages': unread_messages,
        'total_revenue': total_revenue,
        'recent_orders': recent_orders,
        'recent_messages': recent_messages,
        'low_stock_products': low_stock_products,
        'chart_labels': json.dumps(chart_labels),
        'chart_revenue': json.dumps(chart_revenue),
        'chart_count': json.dumps(chart_count),
        'page_title': 'Tableau de Bord',
    }
    return render(request, 'admin_custom/dashboard.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_export_orders(request):
    """Exporte les commandes au format CSV."""
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="Rapport_Commandes_STANTECH_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['N° Commande', 'Client', 'Email', 'Téléphone', 'Adresse', 'Ville', 'Montant (FCFA)', 'Statut', 'Date'])

    orders = Order.objects.all().order_by('-created_at')
    status_labels = dict(Order.STATUS_CHOICES)

    for o in orders:
        writer.writerow([
            o.order_number,
            o.customer_name,
            o.customer_email,
            o.customer_phone,
            o.delivery_address,
            o.city,
            int(o.total_amount),
            status_labels.get(o.status, o.status),
            o.created_at.strftime('%d/%m/%Y %H:%M')
        ])

    return response



# ============================================================
# PRODUCTS (BOUTIQUE)
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_products(request):
    search = request.GET.get('q', '')
    category_filter = request.GET.get('cat', '')
    products = Product.objects.select_related('category').all()
    if search:
        products = products.filter(Q(name__icontains=search) | Q(short_description__icontains=search))
    if category_filter:
        products = products.filter(category__slug=category_filter)
    categories = Category.objects.filter(type='product')
    context = {
        'products': products,
        'categories': categories,
        'search': search,
        'category_filter': category_filter,
        'page_title': 'Gestion Boutique',
    }
    return render(request, 'admin_custom/products.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_product_add(request):
    categories = Category.objects.filter(type='product')
    if request.method == 'POST':
        try:
            name = request.POST['name']
            price = request.POST['price']
            discount_price = request.POST.get('discount_price') or None
            stock = request.POST.get('stock', 10)
            badge = request.POST.get('badge', '')
            short_description = request.POST['short_description']
            full_description = request.POST.get('full_description', '')
            image_url = request.POST.get('image_url', '')
            image_file = request.FILES.get('image_file')
            if image_file:
                fs = FileSystemStorage()
                filename = fs.save(f"products/{image_file.name}", image_file)
                image_url = fs.url(filename)
            specs = request.POST.get('specs', '')
            new_category_name = request.POST.get('new_category_name', '').strip()
            if new_category_name:
                cat, _ = Category.objects.get_or_create(name=new_category_name, type='product')
                category_id = cat.id
            else:
                category_id = request.POST.get('category') or None
            is_featured = request.POST.get('is_featured') == 'on'
            is_active = request.POST.get('is_active') == 'on'

            product = Product.objects.create(
                name=name,
                price=price,
                discount_price=discount_price,
                stock=stock,
                badge=badge,
                short_description=short_description,
                full_description=full_description,
                image_url=image_url,
                specs=specs,
                category_id=category_id,
                is_featured=is_featured,
                is_active=is_active,
            )
            messages.success(request, f'✅ Produit "{product.name}" ajouté avec succès !')
            return redirect('dashboard_products')
        except Exception as e:
            messages.error(request, f'❌ Erreur : {e}')

    context = {'categories': categories, 'page_title': 'Ajouter un Produit'}
    return render(request, 'admin_custom/product_form.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.filter(type='product')
    if request.method == 'POST':
        try:
            product.name = request.POST['name']
            product.price = request.POST['price']
            product.discount_price = request.POST.get('discount_price') or None
            product.stock = request.POST.get('stock', 10)
            product.badge = request.POST.get('badge', '')
            product.short_description = request.POST['short_description']
            product.full_description = request.POST.get('full_description', '')
            product.image_url = request.POST.get('image_url', '')
            image_file = request.FILES.get('image_file')
            if image_file:
                fs = FileSystemStorage()
                filename = fs.save(f"products/{image_file.name}", image_file)
                product.image_url = fs.url(filename)
            product.specs = request.POST.get('specs', '')
            new_category_name = request.POST.get('new_category_name', '').strip()
            if new_category_name:
                cat, _ = Category.objects.get_or_create(name=new_category_name, type='product')
                product.category_id = cat.id
            else:
                product.category_id = request.POST.get('category') or None
            product.is_featured = request.POST.get('is_featured') == 'on'
            product.is_active = request.POST.get('is_active') == 'on'
            product.slug = ''  # reset slug so it regenerates
            product.save()
            messages.success(request, f'✅ Produit "{product.name}" mis à jour !')
            return redirect('dashboard_products')
        except Exception as e:
            messages.error(request, f'❌ Erreur : {e}')

    context = {'product': product, 'categories': categories, 'page_title': 'Modifier le Produit'}
    return render(request, 'admin_custom/product_form.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        name = product.name
        product.delete()
        messages.success(request, f'🗑️ Produit "{name}" supprimé.')
    return redirect('dashboard_products')


# ============================================================
# SERVICES
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_services(request):
    services = Service.objects.select_related('category').all()
    context = {'services': services, 'page_title': 'Gestion des Services'}
    return render(request, 'admin_custom/services.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_service_add(request):
    categories = Category.objects.filter(type='service')
    if request.method == 'POST':
        try:
            service = Service.objects.create(
                title=request.POST['title'],
                icon=request.POST.get('icon', 'cpu'),
                short_description=request.POST['short_description'],
                full_description=request.POST.get('full_description', ''),
                features=request.POST.get('features', ''),
                image_url=request.POST.get('image_url', ''),
                category_id=request.POST.get('category') or None,
                is_featured=request.POST.get('is_featured') == 'on',
                is_active=request.POST.get('is_active') == 'on',
                order=request.POST.get('order', 0),
            )
            messages.success(request, f'✅ Service "{service.title}" ajouté avec succès !')
            return redirect('dashboard_services')
        except Exception as e:
            messages.error(request, f'❌ Erreur : {e}')

    context = {'categories': categories, 'page_title': 'Ajouter un Service'}
    return render(request, 'admin_custom/service_form.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_service_edit(request, pk):
    service = get_object_or_404(Service, pk=pk)
    categories = Category.objects.filter(type='service')
    if request.method == 'POST':
        try:
            service.title = request.POST['title']
            service.icon = request.POST.get('icon', 'cpu')
            service.short_description = request.POST['short_description']
            service.full_description = request.POST.get('full_description', '')
            service.features = request.POST.get('features', '')
            service.image_url = request.POST.get('image_url', '')
            service.category_id = request.POST.get('category') or None
            service.is_featured = request.POST.get('is_featured') == 'on'
            service.is_active = request.POST.get('is_active') == 'on'
            service.order = request.POST.get('order', 0)
            service.slug = ''
            service.save()
            messages.success(request, f'✅ Service "{service.title}" mis à jour !')
            return redirect('dashboard_services')
        except Exception as e:
            messages.error(request, f'❌ Erreur : {e}')

    context = {'service': service, 'categories': categories, 'page_title': 'Modifier le Service'}
    return render(request, 'admin_custom/service_form.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_service_delete(request, pk):
    service = get_object_or_404(Service, pk=pk)
    if request.method == 'POST':
        name = service.title
        service.delete()
        messages.success(request, f'🗑️ Service "{name}" supprimé.')
    return redirect('dashboard_services')


# ============================================================
# REALISATIONS
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_realisations(request):
    realizations = Realization.objects.select_related('category').all()
    context = {'realizations': realizations, 'page_title': 'Gestion des Réalisations'}
    return render(request, 'admin_custom/realisations.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_realisation_add(request):
    categories = Category.objects.filter(type='realization')
    if request.method == 'POST':
        try:
            realization = Realization.objects.create(
                title=request.POST['title'],
                client_name=request.POST.get('client_name', ''),
                short_description=request.POST['short_description'],
                full_description=request.POST.get('full_description', ''),
                tech_stack=request.POST.get('tech_stack', ''),
                image_url=request.POST.get('image_url', ''),
                project_url=request.POST.get('project_url', ''),
                completion_date=request.POST.get('completion_date') or None,
                category_id=request.POST.get('category') or None,
                is_featured=request.POST.get('is_featured') == 'on',
                is_active=request.POST.get('is_active') == 'on',
            )
            messages.success(request, f'✅ Réalisation "{realization.title}" ajoutée !')
            return redirect('dashboard_realisations')
        except Exception as e:
            messages.error(request, f'❌ Erreur : {e}')

    context = {'categories': categories, 'page_title': 'Ajouter une Réalisation'}
    return render(request, 'admin_custom/realisation_form.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_realisation_edit(request, pk):
    realization = get_object_or_404(Realization, pk=pk)
    categories = Category.objects.filter(type='realization')
    if request.method == 'POST':
        try:
            realization.title = request.POST['title']
            realization.client_name = request.POST.get('client_name', '')
            realization.short_description = request.POST['short_description']
            realization.full_description = request.POST.get('full_description', '')
            realization.tech_stack = request.POST.get('tech_stack', '')
            realization.image_url = request.POST.get('image_url', '')
            realization.project_url = request.POST.get('project_url', '')
            realization.completion_date = request.POST.get('completion_date') or None
            realization.category_id = request.POST.get('category') or None
            realization.is_featured = request.POST.get('is_featured') == 'on'
            realization.is_active = request.POST.get('is_active') == 'on'
            realization.slug = ''
            realization.save()
            messages.success(request, f'✅ Réalisation "{realization.title}" mise à jour !')
            return redirect('dashboard_realisations')
        except Exception as e:
            messages.error(request, f'❌ Erreur : {e}')

    context = {'realization': realization, 'categories': categories, 'page_title': 'Modifier la Réalisation'}
    return render(request, 'admin_custom/realisation_form.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_realisation_delete(request, pk):
    realization = get_object_or_404(Realization, pk=pk)
    if request.method == 'POST':
        name = realization.title
        realization.delete()
        messages.success(request, f'🗑️ Réalisation "{name}" supprimée.')
    return redirect('dashboard_realisations')


# ============================================================
# ORDERS
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_orders(request):
    status_filter = request.GET.get('status', '')
    orders = Order.objects.prefetch_related('items').all()
    if status_filter:
        orders = orders.filter(status=status_filter)
    context = {
        'orders': orders,
        'status_filter': status_filter,
        'status_choices': Order.STATUS_CHOICES,
        'page_title': 'Gestion des Commandes',
    }
    return render(request, 'admin_custom/orders.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status:
            order.status = new_status
            order.save()
            messages.success(request, f'Statut mis à jour : {order.get_status_display()}')
            return redirect('dashboard_order_detail', pk=pk)
    context = {
        'order': order,
        'status_choices': Order.STATUS_CHOICES,
        'page_title': f'Commande #{order.order_number}',
    }
    return render(request, 'admin_custom/order_detail.html', context)


# ============================================================
# MESSAGES
# ============================================================

@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_messages(request):
    msgs = ContactMessage.objects.all()
    context = {'msgs': msgs, 'page_title': 'Messages de Contact'}
    return render(request, 'admin_custom/messages.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_message_read(request, pk):
    msg = get_object_or_404(ContactMessage, pk=pk)
    msg.is_read = True
    msg.save()
    context = {'msg': msg, 'page_title': 'Lire le Message'}
    return render(request, 'admin_custom/message_detail.html', context)


@login_required(login_url='/dashboard/login/')
@user_passes_test(is_superuser, login_url='/dashboard/login/')
def dashboard_message_delete(request, pk):
    msg = get_object_or_404(ContactMessage, pk=pk)
    if request.method == 'POST':
        msg.delete()
        messages.success(request, 'Message supprimé.')
    return redirect('dashboard_messages')
